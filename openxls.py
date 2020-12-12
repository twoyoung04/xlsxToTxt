#-*- coding:UTF-8 -*-

import os
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Fill, Alignment, PatternFill

def write_new_excel(file_name):
    # 创建一个excel文档
    wb = Workbook()
    # 获得当前激活的sheet对象
    ws = wb.active
    # 给A2单元格赋值
    ws['A2'] = 'This is A2 cell'
    # 一行添加多列数据
    ws.append([1, 2, 'hello'])
    # 添加新的sheet
    ws = wb.create_sheet(title='NewInfo',index=0)
    # 设置单元格的值
    ws['A1'] = 'This is new sheet'

    # 保存excel
    wb.save(file_name)


'''
 def read_update_excel(file_name):
    # 加载Excel表
    wb = load_workbook(file_name)
    # 打印sheet数量
    print('sheet count:', len(wb.sheetnames))
    # 打印所有sheet名字
    print('sheet name list:', wb.sheetnames)
    # 获取第一个sheet对象
    ws = wb[wb.sheetnames[0]]
    # 打印sheet表行数和列数
    print('rows count:', ws.max_row, 'cols count:', ws.max_column)
    # 更新单元格A1的内容
    ws['A1'] = 'this is A1'
    # 在第二行位置插入一行
    ws.insert_rows(2)
    # 删除第五行
    ws.delete_rows(5)
    # 获取单元格对象，对应B2单元格
    cell = ws.cell(2,2)
    # 设置单元格内容
    cell.value = 'this is B2'
    # 修改字体格式为粗体
    cell.font = Font(bold=True)
    # 修改单元格格式
    cell.fill = PatternFill("solid", fgColor="F0CDCD")

    # 保存原文件或另存一个文件
    wb.save(file_name)
'''

def get_path_list(dir_path):
    return os.listdir(dir_path)

def check_xlsx(file_path):
    try:
        ext = file_path.rsplit('.')[1]
        return ext == 'xlsx' or ext == 'xls'
    except:
        print('file has no extention\n')
    return false

def is_number(s):
    try:
        float(s)
        return True
    except ValueError:
        pass
 
    try:
        import unicodedata
        unicodedata.numeric(s)
        return True
    except (TypeError, ValueError):
        pass
 
    return False

def xls_to_txt(file_path, output_path = None):
    # 加载Excel表
    wb = load_workbook(file_path)
     # 获取第一个sheet对象
    ws = wb[wb.sheetnames[0]]

    rows = ws.max_row  # 行数
    cols = ws.max_column  # 列数
    
    print(' start to read')
    for i in range(1, rows + 1):  # 从第二行开始读取
        for j in range(1, cols + 1):
            cellValue = ws.cell(row=i, column=j).value
            if is_number(cellValue):
                print(str(cellValue) + '\t')
            else :
                print(cellValue + '\t   ')
        print('\n')

def xls_to_txt_dir(dir_path):

    path_list = get_path_list(dir_path)
    for file_path in path_list:
        if check_xlsx(file_path):
            xls_to_txt(file_path)

if __name__ == '__main__':
    
    dir_path = './'
    xls_to_txt_dir(dir_path)